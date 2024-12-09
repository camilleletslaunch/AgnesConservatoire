import logging

from openpyxl.pivot.fields import Boolean
from openpyxl.workbook import Workbook
from pandas.core.interchange.dataframe_protocol import DataFrame

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook

## TODO: use this instead: file_path = os.path.join('1 - Semis - Années antérieures', year, f'Semis serre {year}.xls')
## TODO: fix the iloc: line_semis['N° lot stock'].iloc[0]

ORIGINE_LOT_DIRECT = ['Don producteur / particulier', 'Inconnue', 'Institiut technique', 'Jardin botanique', 'Nature', 'Pépiniériste / semencier']
FIRST_PATTERN_TEST = 'G77 bis        11 godets en fleurs (11-ps-531) repiq le 8/06/2011'

def open_fichier_semis_serre(yr: str) -> DataFrame:
    semis_file_full_path = SEMIS_ANNEES_ANTERIEURES_PATH + PATH_SEP + yr + PATH_SEP + 'Semis serre '+yr +'.xls'
    logging.debug("Opening " + semis_file_full_path + " tab SEMIS SERRE")
    return pd.read_excel(semis_file_full_path, sheet_name='SEMIS SERRE')

def open_fichier_semis_ext(yr: str) -> DataFrame:
    year = int(yr)
    next_year = year + 1
    semis_file_full_path = SEMIS_ANNEES_ANTERIEURES_PATH + PATH_SEP + yr + PATH_SEP + 'Semis exterieurs '+yr + '-' + str(next_year) +'.xls'
    logging.debug("Opening " + semis_file_full_path + " tab SEMIS EXT")
    return pd.read_excel(semis_file_full_path, sheet_name='SEMIS EXT')

def get_file_type_from_first_pattern(first_pattern:str) -> str:
    pt = r"(?<=-)[a-zA-Z]{2}(?=-)"
    logging.debug(f'GETTING FILE TYPE FROM {first_pattern}')
    return re.search(pt, first_pattern).group()

def get_first_pattern(input_str: str) -> re.Match:
    pattern = r"\d{2,4}-[a-zA-Z]{2}-\d{2,4}"
    return re.search(pattern, input_str)

def get_year_from_first_pattern(first_pattern: str) -> str:
    yr_pattern = '^[^-]+'
    match = re.match(yr_pattern, first_pattern)
    yr2d = match.group()[-2:]
    length = len(match.group())
    logging.debug(f"found year {yr2d} in {first_pattern} with {yr_pattern}")
    if int(yr2d) > 30:
        logging.debug(f'found year: {yr2d}. Prefixing 19.')
        return '19' + yr2d
    else:
        logging.debug(f'found year: {yr2d}. Prefixing 20.')
        return '20' + yr2d

def load_wb_from_gc(year:str)->Workbook:
    gc_basename = GRANDE_COLLECTION_PATH + PATH_SEP + year + PATH_SEP + GC_FILENAME + ' '+year
    logging.debug(f"OPENING {gc_basename}")
    df_gc = pd.read_excel(gc_basename + '.xls', sheet_name='GC')
    df_gc.to_excel(gc_basename + '.xlsx', index=False)
    return load_workbook(gc_basename + '.xlsx')

def load_wb_from_jpm(year:str)->Workbook:
    jpm_file_full_path = JARDIN_PLANTES_MENACEES_PATH + PATH_SEP + year + PATH_SEP + 'Jardin plantes menacees ' + year
    logging.debug("Opening " + jpm_file_full_path + ".xls tab JPM")
    df_gc = pd.read_excel(jpm_file_full_path + '.xls', sheet_name='JPM')
    df_gc.to_excel(jpm_file_full_path + '.xlsx', index=False)
    return load_workbook(jpm_file_full_path + '.xlsx')


def get_left_cell(line: pd.DataFrame, year = None, gc_type = True) -> Cell:
    if not year:
        year = line['STOBEGINDATEAVAILABLE'].iloc[0].strftime("%Y")
    search_str = line['ARTSORT'].iloc[0] + ' ' + line['ARTSPECIES'].iloc[0]
    if not pd.isna(line['ARTVARIETY'].iloc[0]):
        search_str += f' {line['ARTVARIETY'].iloc[0]}'
    logging.debug(f'searching {search_str}')
    if gc_type:
        wb = load_wb_from_gc(year)
    else:
        wb = load_wb_from_jpm(year)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == search_str:
                return ws.cell(row=cell.row, column=cell.column - 1)
    logging.error(f"No left cell found for year {year} in fin , sheet GC")
    raise Exception(f"No left cell found for year {year} in fin , sheet GC")


def find_jb(left_cell: Cell) -> str:
    jb_pattern = r"JB"
    if re.search(jb_pattern, left_cell.value):
        logging.debug('FOUND JB IN LEFT_CELL')
        return left_cell.value
    return None

def find_orig_pattern(left_cell: Cell) -> str:
    orig_pattern = r"(?i)orig(ine|\.|)(.*)"
    if re.search(orig_pattern, left_cell.value):
        logging.debug(f'FOUND ORIG IN LEFT_CELL: {left_cell.value}')
        return left_cell.value
    return None

def update_flow_file(df:DataFrame, origine_full_text:str, lot) ->DataFrame:
    logging.debug(f'updating the dataframe line lot: {lot} with {origine_full_text} and Jardin botanique')
    logging.warning('the dataframe was not saved in the file yet')
    df.loc[df['STOIDLOTSTOCK'] == lot, 'Détail origine primaire du lot'] = origine_full_text
    df.loc[df['STOIDLOTSTOCK'] == lot, 'Origine primaire'] = 'Jardin botanique'
    return df

def process_step_4(df: DataFrame, lot: int, line:DataFrame, gc_type = True, year=None, test: bool = False, origine_primaire: str = None, origine_lot: str = None, origine_full_text: str = None) -> DataFrame:
    left_cell = get_left_cell(line=line, gc_type=True)
    origine_full_text = left_cell.value
    logging.debug(f"Found match at {left_cell.coordinate}. Left cell: {left_cell.coordinate}, Value: {left_cell.value}")
    jb_found = find_jb(left_cell)
    if jb_found is None:
        orig_found = find_orig_pattern(left_cell)
    if jb_found is not None or orig_found is not None:
        logging.debug(f"updating flow file with {origine_full_text} for lot={lot}")
        df = update_flow_file(df, origine_full_text, lot)
        return df
    first_pattern_match = get_first_pattern(left_cell.value)
    if first_pattern_match:
        first_pattern = first_pattern_match.group()
        ###################### ETAPE 5 ##########################
        logging.debug(f'FOUND FIRST PATTERN IN LEFT_CELL {left_cell.value}')
        yr = get_year_from_first_pattern(left_cell.value)
        file_type = get_file_type_from_first_pattern(origine_full_text)
        logging.debug(f"year = {yr}, file_type={file_type}")
        match file_type:
            case x if x in ['PS','ps']:
                df_semis = open_fichier_semis_serre(yr)
            case x if x in ['PE','pe']:
                df_semis = open_fichier_semis_ext(yr)
        line_semis = df_semis[df_semis['N° semis'].str.lower()==first_pattern.lower()]
        logging.debug(f"find provenance from {line_semis}")
        prov = line_semis['Provenance'].iloc[0]
        logging.debug(f"FOUND PROVENANCE: {prov}")
        try:
            recu_le = line_semis['Reçu le '].iloc[0]
        except KeyError:
            logging.warning(f'failed to get key Reçu le  when opening semis file for {origine_full_text}. Retrying without the end space')
            recu_le = line_semis['Reçu le'].iloc[0]
        if prov != 'CNPMAI':
            df = update_flow_file(df, origine_full_text, lot)
            logging.debug(f'found origin in fichier serre from {origine_full_text}')
        else:
            try:
                new_lot = line_semis['N° lot stock'].iloc[0]
                return update_flow_file(df, new_lot, test, origine_primaire, origine_lot, origine_full_text)
            except KeyError:
                logging.warning(f'no N° lot stock found in semis serre file for {origine_full_text} using reçu le {recu_le} instead')
            if yr != recu_le:
                return process_step_4(df, lot, line, recu_le, test)
            else:
                raise Exception("Num lot non trouvé dans le fichier de semis ni dans Grande Collection {recu_le}, lot = {lot}")

def process_flow(df: pd.DataFrame, lot: int, test: bool = False, origine_primaire: str = None, origine_lot: str = None, origine_full_text: str = None) -> pd.DataFrame:
    line = df[df['STOIDLOTSTOCK'] == lot]
    if line.empty:
        logging.error(f'{lot} not found')
        return df
    origine_primaire = line['Origine primaire'].iloc[0]
    if not pd.isna(origine_primaire):
        logging.debug(f'{lot} : found origine primaire: {origine_primaire} NO FILE UPDATE FOR THIS LINE')
        return df
    origine_lot = line['Origine lot'].iloc[0]
    match origine_lot:
        case x if x in ORIGINE_LOT_DIRECT:
            logging.debug(f'{lot} : found origine lot: {origine_lot} NO FILE UPDATE FOR THIS LINE')
            return df
        case 'CNPMAI - Grande Collection':
            num_semis = line['N° semis'].iloc[0]
            if pd.isna(num_semis):
                ################## etape 4 #####################
                return process_step_4(df = df, lot = lot, line = line, gc_type=True)
            else:# ############### etape 5 #####################
                num_semis_pattern = get_first_pattern(num_semis)
        case 'CNPMAI - JPM':
            return process_step_4(df = df, lot = lot, line = line, gc_type=False)
        case x if x in ['CNPMAI - Abords Accueil', 'CNPMAI - Serre', 'CNPMAI - Cultures', 'CNPMAI - Jardin Thématique', 'CNPMAI - Ligneux', 'Plantes Ombre']:
            return update_flow_file(df, lot, test, "origine non retrouvee", origine_lot, origine_lot)
        case _:
            raise Exception(f'Value for origine lot not handled: {origine_lot}, lot {lot}')


    logging.debug(f'{lot} : END OF FUNCTION, origine lot: {origine_lot} NO FILE UPDATE FOR THIS LINE')
    return df

