import pandas as pd
import re
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX

lot = 15158
origine_lot_direct = ['Don producteur / particulier', 'Inconnue', 'Institiut technique', 'Jardin botanique', 'Nature', 'Pépiniériste / semencier']
# grande_collection_path = 'R:\\SERVICE TECHNIQUE\\2 - RESSOURCES BOTANIQUES\\3 - Plans et inventaires\\1 - Grande collection'
grande_collection_path = '1 - Grande collection'
gc_fname = 'Grande collection ' # Grande collection 2011.xls
windows_systeme = False
pathsep = '/'
if windows_systeme:
    pathsep = '\'

def get_first_pattern(input_str):
    pattern = r"\d{2,4}-[a-zA-Z]{2}-\d{2,4}"
    res = re.match(pattern, input_str)
    return res

df = pd.read_excel('Consultation_SEM_Flow_PourCamille.xlsx', sheet_name='Origine lots', header=1)
result = df[df['STOIDLOTSTOCK'] == lot]
if not result.empty:
    origine_primaire = result['Origine primaire'].iloc[0]
    if pd.isna(origine_primaire):
        if result['Origine lot'].iloc[0] in origine_lot_direct:
            print('Origine lot: ', result['Origine lot'].iloc[0])
            print('PATH: Origine primaire: ', origine_primaire, '\n\tOrigine lot: ', result['Origine lot'].iloc[0])
            quit()
        else:
            num_semis = get_first_pattern(result['N° semis'].iloc[0])
            if num_semis:
                # ############### etape 5 #####################
                print()
            else:
                year = result['STOBEGINDATEAVAILABLE'].iloc[0].strftime("%Y")
                search_str = result['ARTSORT'].iloc[0] + ' ' + result['ARTSPECIES'].iloc[0]
                if pd.isna(result['ARTVARIETY'].iloc[0]):
                    search_str += ' ' + result['ARTVARIETY'].iloc[0]
                gc_basename = grande_collection_path +pathsep+ year+pathsep+gc_fname+year
                df_gc = pd.read_excel(gc_basename+'.xls', sheet_name='GC')
                df_gc.to_excel(gc_basename + '.xlsx', index=False)
                wb = load_workbook(gc_basename + '.xlsx')
                ws = wb.active
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == search_str:
                            left_cell = ws.cell(row=cell.row, column=cell.column - 1)
                            # print(
                                # f"Found match at {cell.coordinate}. Left cell: {left_cell.coordinate}, Value: {left_cell.value}")
                            # ETAPE 4 ################## ETAPE 4  SLIDE 7 ###########################
                            jb_pattern = r"JB"
                            if re.match(jb_pattern, left_cell.value):
                                print("ORIGINE:", left_cell.value)
                                print("FOUND IN FILE: ", gc_fname+'.xls')
                                orig_pattern = r"(?i)orig(ine|\.|)(.*)"
                                orig_match = re.search(orig_pattern, left_cell.value)
                                if orig_match:
                                    result['Détail origine primaire du lot'] = left_cell.value
                                    result['Origine primaire'] = 'Jardin botanique'
                                quit()
                            else: ############# ETAPE 4 SLIDE 8 #################

                            break
                        else:
                            continue
                        break


    else:
        print('Origine primaire: ', origine_primaire)
        quit()
else:
    print("Aucune ligne ou STOIDLOTSTOCK = 15158")
